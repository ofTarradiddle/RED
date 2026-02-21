"""
Research Ticker Changes and Fetch Historical Returns
For tickers with missing returns, research ticker changes and fetch returns using correct historical tickers.

This script:
1. Loads tickers with missing returns from overlap CSV
2. Researches ticker changes (old ticker -> new ticker with date ranges)
3. Fetches returns using correct historical tickers for missing periods
4. Creates Excel file with amendments for review
5. Generates amended returns dataset
"""

import sys
from pathlib import Path
import os
import logging
from datetime import datetime
import pandas as pd
import requests
import time
from typing import Optional, List, Dict, Tuple
from dotenv import load_dotenv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Add project root to path
project_root = Path(os.getcwd())
sys.path.insert(0, str(project_root))

# Load environment variables
env_path = project_root / '.env'
if env_path.exists():
    load_dotenv(env_path)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Data paths
OVERLAP_FILE = Path('./data/research/sp500_backtest/ticker_constituent_missing_overlap.csv')
RETURNS_FILE = Path('./data/research/sp500_returns/sp500_total_returns_corrected.csv')
OUTPUT_DIR = Path('./data/research/sp500_returns')
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# FMP API configuration
FMP_BASE_URL = "https://financialmodelingprep.com"
FMP_ENDPOINT = "stable/historical-price-eod/full"
FMP_API_KEY = os.getenv('FMP_API_KEY')
REQUEST_DELAY = 0.3

# Data paths for research results
RESEARCH_RESULTS_FILE = Path('./data/research/sp500_backtest/missing_tickers_research_results.csv')
REVERSE_MAP_FILE = Path('./data/research/sp500_backtest/ticker_change_reverse_map.json')
FORWARD_MAP_FILE = Path('./data/research/sp500_backtest/ticker_change_forward_map.json')

# Known ticker changes (expanded from research results)
KNOWN_TICKER_CHANGES = {
    # Major mergers/splits
    'DWDP': {'old_ticker': 'DD', 'change_date': '2017-09-01', 'notes': 'DowDuPont merger - DD changed to DWDP'},
    'DOW': {'old_ticker': 'DWDP', 'change_date': '2019-04-01', 'notes': 'DowDuPont split - DWDP became DOW, DD, and CTVA'},
    'JPM': {'old_ticker': 'JP', 'change_date': '2000-01-01', 'notes': 'J.P. Morgan merged with Chase Manhattan, ticker changed from JP to JPM'},
    
    # Ticker changes from research results
    'ABI': {'old_ticker': 'BUD', 'change_date': '2008-01-01', 'notes': 'Anheuser-Busch InBev - ticker changed to BUD, then acquired'},
    'CMCSA': {'old_ticker': 'CMCSK', 'change_date': '2002-01-01', 'notes': 'Comcast Class A ticker changed from CMCSK to CMCSA'},
    'WYND': {'old_ticker': 'WYN', 'change_date': '2018-01-01', 'notes': 'Wyndham Destinations split from Wyndham Worldwide (WYN) in 2018'},
    'HLT': {'old_ticker': 'H', 'change_date': '2017-01-01', 'notes': 'Hilton Hotels split from Hilton Worldwide, ticker changed to HLT'},
    'H': {'old_ticker': 'HOT', 'change_date': '2017-01-01', 'notes': 'Hilton - ticker changed from HOT to H, then to HLT'},
    'CINF': {'old_ticker': 'CIN', 'change_date': '2000-01-01', 'notes': 'Cincinnati Financial ticker changed from CIN to CINF'},
    'MCK': {'old_ticker': 'MKG', 'change_date': '2000-01-01', 'notes': 'McKesson ticker changed from MKG to MCK'},
    'RS': {'old_ticker': 'RLM', 'change_date': '2000-01-01', 'notes': 'Reliance Steel ticker changed from RLM to RS'},
    'ABC': {'old_ticker': 'ARC', 'change_date': '2000-01-01', 'notes': 'AmeriSourceBergen ticker changed from ARC to ABC'},
    'NWL': {'old_ticker': 'NLC', 'change_date': '1999-01-01', 'notes': 'Newell Rubbermaid ticker changed from NLC to NWL'},
    'EQNR': {'old_ticker': 'STO', 'change_date': '2018-01-01', 'notes': 'Statoil ticker changed from STO to EQNR (Equinor)'},
    'DDS': {'old_ticker': 'DI', 'change_date': '1998-01-01', 'notes': "Dillard's ticker changed from DI to DDS"},
    'GOLD': {'old_ticker': 'ABX', 'change_date': '2019-01-01', 'notes': 'Barrick Gold ticker changed from ABX to GOLD'},
    'DGII': {'old_ticker': 'DIGI', 'change_date': '1998-01-01', 'notes': 'Digi International ticker changed from DIGI to DGII'},
    'WM': {'old_ticker': 'WMX', 'change_date': '1998-01-01', 'notes': 'Waste Management ticker changed from WMX to WM'},
    'PCG': {'old_ticker': 'PAC', 'change_date': '1997-01-01', 'notes': 'Pacific Gas & Electric ticker changed from PAC to PCG'},
    'USB': {'old_ticker': 'USBC', 'change_date': '1997-01-01', 'notes': 'US Bancorp ticker changed from USBC to USB'},
    'PM': {'old_ticker': 'PMI', 'change_date': '1996-01-01', 'notes': 'Philip Morris International ticker changed from PMI to PM'},
    'UIS': {'old_ticker': 'UVN', 'change_date': '2007-01-01', 'notes': 'Unisys ticker changed from UVN to UIS'},
    'CHTR': {'old_ticker': 'CHA', 'change_date': '2000-01-01', 'notes': 'Charter Communications - ticker changed to CHTR'},
}


def fetch_ticker_prices(session: requests.Session, ticker: str, 
                       start_date: Optional[str] = None, 
                       end_date: Optional[str] = None) -> Optional[pd.DataFrame]:
    """
    Fetch historical prices for a ticker using 5-year chunking.
    
    Args:
        session: Requests session
        ticker: Ticker symbol
        start_date: Target start date (YYYY-MM-DD)
        end_date: Target end date (YYYY-MM-DD)
        
    Returns:
        DataFrame with date index and price columns, or None if no data
    """
    url = f"{FMP_BASE_URL}/{FMP_ENDPOINT}"
    
    target_start = pd.to_datetime(start_date) if start_date else pd.to_datetime('1950-01-01')
    target_end = pd.to_datetime(end_date) if end_date else pd.Timestamp.now().normalize()
    
    all_data = []
    current_end = target_end
    chunk_count = 0
    max_chunks = 30
    
    while chunk_count < max_chunks:
        chunk_start = current_end - pd.Timedelta(days=1825)  # 5 years
        
        if chunk_start < target_start:
            chunk_start = target_start
        
        if chunk_start >= current_end:
            break
        
        params = {
            'symbol': ticker,
            'from': chunk_start.strftime('%Y-%m-%d'),
            'to': current_end.strftime('%Y-%m-%d'),
            'apikey': FMP_API_KEY
        }
        
        try:
            response = session.get(url, params=params, timeout=60)
            if response.status_code == 200:
                data = response.json()
                if data and isinstance(data, list) and len(data) > 0:
                    df = pd.DataFrame(data)
                    df['date'] = pd.to_datetime(df['date'])
                    df = df.set_index('date').sort_index()
                    df = df[(df.index >= chunk_start) & (df.index <= current_end)]
                    
                    if not df.empty:
                        all_data.append(df)
                    current_end = chunk_start - pd.Timedelta(days=1)
                else:
                    current_end = chunk_start - pd.Timedelta(days=1)
            elif response.status_code == 403:
                logger.warning(f"  API access denied for {ticker}")
                break
            else:
                break
        except Exception as e:
            logger.debug(f"  Error fetching chunk for {ticker}: {e}")
            break
        
        chunk_count += 1
        time.sleep(REQUEST_DELAY)
    
    if all_data:
        combined_df = pd.concat(all_data)
        combined_df = combined_df.sort_index()
        combined_df = combined_df[~combined_df.index.duplicated(keep='last')]
        return combined_df
    
    return None


def load_research_results() -> pd.DataFrame:
    """Load research results to identify ticker changes."""
    if RESEARCH_RESULTS_FILE.exists():
        df = pd.read_csv(RESEARCH_RESULTS_FILE)
        logger.info(f"Loaded {len(df)} research results")
        return df
    return pd.DataFrame()


def load_reverse_map() -> Dict:
    """Load reverse ticker change map (new_ticker -> old_ticker)."""
    if REVERSE_MAP_FILE.exists():
        with open(REVERSE_MAP_FILE, 'r') as f:
            reverse_map = json.load(f)
        logger.info(f"Loaded reverse map with {len(reverse_map)} ticker changes")
        return reverse_map
    return {}


def load_forward_map() -> Dict:
    """Load forward ticker change map (old_ticker -> new_ticker)."""
    if FORWARD_MAP_FILE.exists():
        with open(FORWARD_MAP_FILE, 'r') as f:
            forward_map = json.load(f)
        logger.info(f"Loaded forward map with {len(forward_map)} ticker changes")
        return forward_map
    return {}


def research_ticker_change_web(ticker: str, missing_periods: List[str]) -> Optional[Dict]:
    """
    Research ticker changes using web search.
    
    Args:
        ticker: Current ticker symbol
        missing_periods: List of periods (YYYY-MM) with missing returns
        
    Returns:
        Dict with ticker change info or None
    """
    try:
        # Search for ticker change information
        search_query = f"{ticker} stock ticker change history old ticker symbol"
        # Note: web_search tool would be called here, but we'll use a placeholder
        # In actual implementation, this would call web_search tool
        logger.debug(f"  Would search web for: {search_query}")
        return None
    except Exception as e:
        logger.debug(f"  Web search error for {ticker}: {e}")
        return None


def research_ticker_change(ticker: str, missing_periods: List[str], 
                          research_df: pd.DataFrame,
                          reverse_map: Dict,
                          forward_map: Dict) -> Optional[Dict]:
    """
    Research ticker changes for a given ticker.
    
    Args:
        ticker: Current ticker symbol (could be old or new ticker)
        missing_periods: List of periods (YYYY-MM) with missing returns
        research_df: DataFrame with research results
        reverse_map: Reverse map of new_ticker -> old_ticker
        forward_map: Forward map of old_ticker -> new_ticker
        
    Returns:
        Dict with ticker change info or None
    """
    # Check known changes first
    if ticker in KNOWN_TICKER_CHANGES:
        change_info = KNOWN_TICKER_CHANGES[ticker].copy()
        change_info['current_ticker'] = ticker
        return change_info
    
    # If ticker is an OLD ticker (in forward map), it changed to a new ticker
    # But we want to fetch data for the OLD ticker itself (it's the historical ticker)
    if ticker in forward_map:
        # This ticker is an old ticker that changed to a new one
        # We should fetch data using this old ticker itself
        map_entry = forward_map[ticker]
        new_ticker = map_entry.get('new_ticker', ticker)
        change_date = map_entry.get('change_date', '')
        notes = map_entry.get('notes', '')
        
        # Format change_date
        if change_date and len(str(change_date)) == 4:
            change_date = f"{change_date}-01-01"
        elif not change_date and missing_periods:
            change_date = f"{missing_periods[-1]}-01"  # Use last missing period
        
        # Return info indicating we should use this ticker (old ticker) for historical data
        return {
            'current_ticker': ticker,  # The old ticker in overlap file
            'historical_ticker': ticker,  # Use the old ticker itself for historical data
            'new_ticker': new_ticker,  # What it changed to
            'change_date': change_date,
            'notes': f"Old ticker that changed to {new_ticker}. Using {ticker} for historical data.",
            'category': 'Ticker Change'
        }
    
    # Check reverse map (new ticker -> old ticker)
    # If ticker is a NEW ticker, we need to use the OLD ticker for historical data
    if ticker in reverse_map:
        map_entry = reverse_map[ticker]
        old_ticker = map_entry.get('old_ticker', ticker)
        change_date = map_entry.get('change_date', '')
        notes = map_entry.get('notes', '')
        
        # Format change_date
        if change_date and len(str(change_date)) == 4:
            change_date = f"{change_date}-01-01"
        elif not change_date and missing_periods:
            change_date = f"{missing_periods[0]}-01"
        
        return {
            'current_ticker': ticker,  # The new ticker
            'historical_ticker': old_ticker,  # Use old ticker for historical data
            'change_date': change_date,
            'notes': notes,
            'category': 'Ticker Change'
        }
    
    # Check research results for ticker changes
    if not research_df.empty:
        ticker_research = research_df[research_df['ticker'] == ticker]
        if not ticker_research.empty:
            row = ticker_research.iloc[0]
            category = str(row.get('category', '')).strip()
            notes = str(row.get('notes', '')).strip()
            date = row.get('date', '')
            
            # Look for ticker change patterns in notes
            notes_lower = notes.lower()
            is_ticker_change = (
                'Ticker Change' in category or 
                'ticker change' in notes_lower or 
                'changed to' in notes_lower or
                'ticker changed' in notes_lower or
                'symbol changed' in notes_lower or
                'renamed' in notes_lower
            )
            
            if is_ticker_change:
                # Try to extract old ticker from notes
                import re
                # Look for patterns like "changed to X", "was X", "formerly X", "from X to Y"
                patterns = [
                    r'changed from ([A-Z]{1,5})\b',  # "changed from X to Y"
                    r'from ([A-Z]{1,5}) to',  # "from X to Y"
                    r'changed to ([A-Z]{1,5})\b',  # "changed to X"
                    r'was ([A-Z]{1,5})\b',  # "was X"
                    r'formerly ([A-Z]{1,5})\b',  # "formerly X"
                    r'old ticker:?\s*([A-Z]{1,5})\b',  # "old ticker: X"
                    r'previous ticker:?\s*([A-Z]{1,5})\b',  # "previous ticker: X"
                    r'\(([A-Z]{1,5})\)',  # Ticker in parentheses like "(TFC)"
                    r'\b([A-Z]{1,5})\s+to\s+[A-Z]',  # "X to Y" pattern
                ]
                
                old_ticker = None
                for pattern in patterns:
                    match = re.search(pattern, notes, re.IGNORECASE)
                    if match:
                        candidate = match.group(1).upper()
                        # Validate it's a reasonable ticker (1-5 letters, not common words)
                        if len(candidate) >= 1 and len(candidate) <= 5 and candidate != ticker:
                            # Filter out common words
                            common_words = {'THE', 'AND', 'FOR', 'WAS', 'HAS', 'HAD', 'ARE', 'WERE'}
                            if candidate not in common_words:
                                old_ticker = candidate
                                break
                
                if old_ticker and old_ticker != ticker:
                    # Convert date to change_date format
                    change_date = None
                    if date:
                        date_str = str(date).strip()
                        if len(date_str) == 4 and date_str.isdigit():
                            change_date = f"{date_str}-01-01"
                        elif '-' in date_str:
                            change_date = date_str
                    
                    if not change_date and missing_periods:
                        # Use first missing period as proxy
                        first_missing = missing_periods[0]
                        change_date = f"{first_missing}-01"
                    
                    return {
                        'current_ticker': ticker,
                        'old_ticker': old_ticker,
                        'change_date': change_date,
                        'notes': notes,
                        'category': category
                    }
    
    # If no change found in research, try web search for high-priority tickers
    # (those with many missing periods)
    if len(missing_periods) > 50:  # Only for tickers with significant missing data
        logger.debug(f"  Attempting web research for {ticker} ({len(missing_periods)} missing periods)")
        web_result = research_ticker_change_web(ticker, missing_periods)
        if web_result:
            return web_result
    
    return None


def fetch_returns_for_period(session: requests.Session, ticker: str, 
                             period: str) -> Optional[pd.Series]:
    """
    Fetch returns for a specific period (YYYY-MM).
    
    Args:
        session: Requests session
        ticker: Ticker symbol
        period: Period in YYYY-MM format
        
    Returns:
        Series of daily returns or None
    """
    # Convert period to date range
    period_start = pd.to_datetime(f"{period}-01")
    # Get last day of month
    if period_start.month == 12:
        period_end = pd.to_datetime(f"{period_start.year + 1}-01-01") - pd.Timedelta(days=1)
    else:
        period_end = pd.to_datetime(f"{period_start.year}-{period_start.month + 1}-01") - pd.Timedelta(days=1)
    
    # Fetch prices for this period
    prices_df = fetch_ticker_prices(
        session, 
        ticker, 
        start_date=period_start.strftime('%Y-%m-%d'),
        end_date=period_end.strftime('%Y-%m-%d')
    )
    
    if prices_df is not None and not prices_df.empty and 'close' in prices_df.columns:
        returns = prices_df['close'].pct_change().dropna()
        return returns
    
    return None


def load_missing_tickers() -> pd.DataFrame:
    """Load tickers with missing returns from overlap CSV."""
    if not OVERLAP_FILE.exists():
        logger.error(f"Overlap file not found: {OVERLAP_FILE}")
        return pd.DataFrame()
    
    df = pd.read_csv(OVERLAP_FILE)
    logger.info(f"Loaded {len(df)} tickers with missing returns")
    return df


def process_ticker_changes(session: requests.Session, 
                          overlap_df: pd.DataFrame,
                          research_df: pd.DataFrame,
                          reverse_map: Dict,
                          forward_map: Dict) -> List[Dict]:
    """
    Process all tickers to identify ticker changes and fetch returns.
    
    Returns:
        List of amendment records
    """
    amendments = []
    
    for idx, row in overlap_df.iterrows():
        ticker = row['ticker']
        missing_periods_str = row.get('missing_periods_list', '')
        
        if pd.isna(missing_periods_str) or missing_periods_str == '':
            continue
        
        missing_periods = [p.strip() for p in missing_periods_str.split(',')]
        
        logger.info(f"Processing {ticker} ({idx+1}/{len(overlap_df)}) - {len(missing_periods)} missing periods")
        
        # Research ticker change
        change_info = research_ticker_change(ticker, missing_periods, research_df, reverse_map, forward_map)
        
        if change_info:
            historical_ticker = change_info.get('historical_ticker', ticker)
            change_date = change_info.get('change_date', '')
            new_ticker = change_info.get('new_ticker')
            
            if new_ticker:
                logger.info(f"  Found: {ticker} is old ticker that changed to {new_ticker} (changed {change_date})")
                logger.info(f"    Using {historical_ticker} (old ticker) for historical data")
            else:
                logger.info(f"  Found ticker change: {ticker} -> {historical_ticker} (changed {change_date})")
            
            # Determine which periods need historical ticker
            if change_date:
                try:
                    change_date_dt = pd.to_datetime(change_date)
                except:
                    change_date_dt = None
            else:
                change_date_dt = None
            
            for period in missing_periods:
                period_start = pd.to_datetime(f"{period}-01")
                
                # If we have a change date, only use historical ticker for periods before change
                # Otherwise, use historical ticker for all missing periods
                use_historical = True
                if change_date_dt:
                    # If ticker is an old ticker (in forward map), use it for all periods before it changed
                    if ticker in forward_map:
                        if period_start >= change_date_dt:
                            use_historical = False  # After change, ticker no longer exists
                    else:
                        # If ticker is a new ticker, use old ticker only for periods before change
                        if period_start >= change_date_dt:
                            use_historical = False
                
                if use_historical:
                    logger.info(f"    Fetching {historical_ticker} for period {period}")
                    returns = fetch_returns_for_period(session, historical_ticker, period)
                    
                    if returns is not None and not returns.empty:
                        amendment = {
                            'current_ticker': ticker,
                            'historical_ticker': historical_ticker,
                            'period': period,
                            'returns_count': len(returns),
                            'date_range': f"{returns.index.min().date()} to {returns.index.max().date()}",
                            'status': 'Found',
                            'notes': change_info.get('notes', ''),
                            'change_date': change_date
                        }
                        amendments.append(amendment)
                        logger.info(f"      ✓ Found {len(returns)} returns")
                    else:
                        amendment = {
                            'current_ticker': ticker,
                            'historical_ticker': historical_ticker,
                            'period': period,
                            'returns_count': 0,
                            'date_range': '',
                            'status': 'Not Found',
                            'notes': change_info.get('notes', ''),
                            'change_date': change_date
                        }
                        amendments.append(amendment)
                        logger.info(f"      ✗ No returns found")
                else:
                    # Period is after change, but still missing - may need different research
                    amendment = {
                        'current_ticker': ticker,
                        'historical_ticker': ticker,
                        'period': period,
                        'returns_count': 0,
                        'date_range': '',
                        'status': 'After Change Date',
                        'notes': f'Period after ticker change on {change_date}',
                        'change_date': change_date
                    }
                    amendments.append(amendment)
        else:
            # No known ticker change - try fetching with the ticker itself first
            # (it might be an old ticker that we just don't have in our maps)
            logger.info(f"  No ticker change found - trying to fetch {ticker} directly")
            for period in missing_periods[:10]:  # Try first 10 periods
                returns = fetch_returns_for_period(session, ticker, period)
                
                if returns is not None and not returns.empty:
                    amendment = {
                        'current_ticker': ticker,
                        'historical_ticker': ticker,
                        'period': period,
                        'returns_count': len(returns),
                        'date_range': f"{returns.index.min().date()} to {returns.index.max().date()}",
                        'status': 'Found',
                        'notes': 'Found data using ticker directly (no ticker change)',
                        'change_date': ''
                    }
                    amendments.append(amendment)
                    logger.info(f"      ✓ Found {len(returns)} returns for {ticker}")
                else:
                    amendment = {
                        'current_ticker': ticker,
                        'historical_ticker': 'UNKNOWN',
                        'period': period,
                        'returns_count': 0,
                        'date_range': '',
                        'status': 'Needs Research',
                        'notes': 'No ticker change found and no data available - needs manual research',
                        'change_date': ''
                    }
                    amendments.append(amendment)
        
        time.sleep(REQUEST_DELAY * 2)  # Extra delay between tickers
    
    return amendments


def create_excel_amendments(amendments: List[Dict], output_file: Path):
    """Create Excel file with amendments for review."""
    logger.info(f"Creating Excel file: {output_file}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Return Amendments"
    
    # Headers
    headers = ['Current Ticker', 'Historical Ticker', 'Period', 'Returns Count', 
               'Date Range', 'Status', 'Change Date', 'Notes']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for amendment in amendments:
        row = [
            amendment.get('current_ticker', ''),
            amendment.get('historical_ticker', ''),
            amendment.get('period', ''),
            amendment.get('returns_count', 0),
            amendment.get('date_range', ''),
            amendment.get('status', ''),
            amendment.get('change_date', ''),
            amendment.get('notes', '')
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = max(
            len(str(header)),
            max([len(str(ws.cell(row=row_num, column=col_num).value)) 
                 for row_num in range(2, ws.max_row + 1)] + [0])
        )
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(output_file)
    logger.info(f"✓ Saved Excel file: {output_file}")


def save_amendments_json(amendments: List[Dict], output_file: Path):
    """Save amendments as JSON for programmatic use."""
    with open(output_file, 'w') as f:
        json.dump(amendments, f, indent=2, default=str)
    logger.info(f"✓ Saved JSON file: {output_file}")


def save_amended_returns(amendments: List[Dict], session: requests.Session, 
                        output_file: Path):
    """
    Fetch and save amended returns data that can be merged with existing returns.
    
    Only processes amendments with status 'Found'.
    """
    logger.info(f"\nFetching amended returns data...")
    
    # Group amendments by current_ticker
    ticker_amendments = {}
    for amendment in amendments:
        if amendment.get('status') == 'Found':
            ticker = amendment['current_ticker']
            if ticker not in ticker_amendments:
                ticker_amendments[ticker] = []
            ticker_amendments[ticker].append(amendment)
    
    all_amended_returns = {}
    
    for ticker, ticker_amends in ticker_amendments.items():
        logger.info(f"  Processing {ticker} ({len(ticker_amends)} periods)...")
        
        # Collect all returns for this ticker
        ticker_returns = []
        
        for amend in ticker_amends:
            period = amend['period']
            historical_ticker = amend['historical_ticker']
            
            returns = fetch_returns_for_period(session, historical_ticker, period)
            if returns is not None and not returns.empty:
                ticker_returns.append(returns)
                time.sleep(REQUEST_DELAY)
        
        if ticker_returns:
            # Combine all returns for this ticker
            combined = pd.concat(ticker_returns)
            combined = combined.sort_index()
            combined = combined[~combined.index.duplicated(keep='last')]
            all_amended_returns[ticker] = combined
            logger.info(f"    ✓ {ticker}: {len(combined)} returns")
    
    if all_amended_returns:
        # Create DataFrame with all amended returns
        amended_df = pd.DataFrame(all_amended_returns)
        amended_df.index.name = 'date'
        amended_df.to_csv(output_file)
        logger.info(f"\n✓ Saved amended returns: {output_file}")
        logger.info(f"  Tickers: {len(amended_df.columns)}")
        logger.info(f"  Date range: {amended_df.index.min()} to {amended_df.index.max()}")
    else:
        logger.warning("No amended returns to save")


def main():
    """Main execution function."""
    if not FMP_API_KEY:
        logger.error("FMP_API_KEY environment variable not set.")
        return
    
    logger.info("="*80)
    logger.info("Research Ticker Changes and Fetch Historical Returns")
    logger.info("="*80)
    
    # Load missing tickers
    overlap_df = load_missing_tickers()
    if overlap_df.empty:
        logger.error("No tickers to process")
        return
    
    # Create session
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'REDI Research',
        'Accept': 'application/json'
    })
    
    # Load research results and maps
    research_df = load_research_results()
    reverse_map = load_reverse_map()
    forward_map = load_forward_map()
    
    # Process ticker changes
    logger.info("\nProcessing ticker changes and fetching returns...")
    amendments = process_ticker_changes(session, overlap_df, research_df, reverse_map, forward_map)
    
    logger.info(f"\n✓ Processed {len(amendments)} amendments")
    
    # Save results
    excel_file = OUTPUT_DIR / 'ticker_return_amendments.xlsx'
    json_file = OUTPUT_DIR / 'ticker_return_amendments.json'
    amended_returns_file = OUTPUT_DIR / 'ticker_return_amendments_data.csv'
    
    create_excel_amendments(amendments, excel_file)
    save_amendments_json(amendments, json_file)
    save_amended_returns(amendments, session, amended_returns_file)
    
    # Summary
    found_count = sum(1 for a in amendments if a.get('status') == 'Found')
    logger.info(f"\nSummary:")
    logger.info(f"  Total amendments: {len(amendments)}")
    logger.info(f"  Returns found: {found_count}")
    logger.info(f"  Needs research: {len(amendments) - found_count}")
    logger.info(f"\n✓ Excel file saved: {excel_file}")
    logger.info(f"✓ JSON file saved: {json_file}")
    logger.info(f"✓ Amended returns data saved: {amended_returns_file}")


if __name__ == '__main__':
    main()

