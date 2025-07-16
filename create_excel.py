import pandas as pd
import numpy as np
from datetime import datetime, timedelta

# Create the data
dates = pd.date_range(start='2024-01-01', end='2024-12-19', freq='D')

# Create more realistic price movements with volatility
np.random.seed(42)  # For reproducible results

# Start with base values
red_etf_base = 100
sp500_base = 100

red_etf = []
nav = []
sp500 = []
premium_discount = []

for i in range(len(dates)):
    # Add some daily variation
    red_etf_change = np.random.normal(0.15, 0.8)  # Daily return with volatility
    sp500_change = np.random.normal(0.08, 0.6)
    
    if i == 0:
        red_etf.append(red_etf_base)
        nav.append(red_etf_base)  # NAV starts same as RED ETF
        sp500.append(sp500_base)
        premium_discount.append(100.00)  # Start at 100 (no premium/discount)
    else:
        red_etf.append(red_etf[-1] * (1 + red_etf_change/100))
        sp500.append(sp500[-1] * (1 + sp500_change/100))
        
        # NAV should be very close to RED ETF price
        # Add tiny random variation to NAV (much smaller than RED ETF variation)
        nav_variation = np.random.normal(0, 0.02)  # Very small variation
        nav.append(red_etf[-1] * (1 + nav_variation/100))
        
        # Calculate premium/discount as the percentage difference
        # Premium/Discount = (RED_ETF / NAV - 1) * 100 + 100
        # This gives us a value around 100 where 100 = no premium/discount
        if nav[-1] != 0:
            premium_pct = (red_etf[-1] / nav[-1] - 1) * 100
            premium_discount.append(100 + premium_pct)
        else:
            premium_discount.append(100.00)

# Create DataFrame
df = pd.DataFrame({
    'Date': dates,
    'RED_ETF': [round(x, 2) for x in red_etf],
    'NAV': [round(x, 2) for x in nav],
    'SP500': [round(x, 2) for x in sp500],
    'Premium_Discount': [round(x, 2) for x in premium_discount]
})

# Save to Excel with formatting
with pd.ExcelWriter('data/performance_data.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Performance Data', index=False)
    
    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Performance Data']
    
    # Format the date column
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    
    for col in range(1, 6):  # Columns A to E
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        worksheet.column_dimensions[column_letter].width = adjusted_width

print("Excel file created: data/performance_data.xlsx")
print(f"Data points: {len(df)}")
print("Sample data:")
print(df.head())
print("\nLatest values:")
print(df.tail())
print(f"\nPremium/Discount range: {df['Premium_Discount'].min():.2f} to {df['Premium_Discount'].max():.2f}")
print(f"Premium/Discount mean: {df['Premium_Discount'].mean():.2f}")
print(f"\nRED ETF vs NAV correlation: {df['RED_ETF'].corr(df['NAV']):.4f}")
print(f"Average difference between RED ETF and NAV: {abs(df['RED_ETF'] - df['NAV']).mean():.4f}") 